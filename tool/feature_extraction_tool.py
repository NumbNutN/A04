####################################################################################
#                                特征提取工具组                                     #
#  -- 用于训练测试数据预处理，特征工程等                                              #
#  -- 通过 import classification_tool as cft 使用 cft.func                          #
####################################################################################

#特征工程 - 文本分词需要使用
import spacy

#读取xlsx文件需要使用
from openpyxl import load_workbook

#读取csv文件需要使用
import csv
import numpy as np

#降维等其他矩阵运算
from sklearn.manifold import TSNE

from tqdm import tqdm

class DataFeature:

    #从项目根路径开始
    dataSetPath:str = "data/"

    
################################################################
#   update on 2023-3-18                                         #
#   数据集train2.csv的排列发生了改动                              #
#   请使用excel对数据集进行一些微小的改动已支持dataFeatureList的变动#
#   使用excel打开train2.csv 选中标签所在列——排序——升序——保存改动    #
##################################################################
    
    dataFeatureList:list = \
    [
        {"fileName":"train2.csv","type":"信贷理财","range":range(2,723)},
        {"fileName":"train2.csv","type":"刷单诈骗","range":range(725,1027)},
        {"fileName":"train2.csv","type":"婚恋交友","range":range(1167,7210)},

        {"fileName":"train3.xlsx","type":"刷单诈骗","range":range(243,430)},
        {"fileName":"train3.xlsx","type":"信贷理财","range":range(1,241)},
        

        {"fileName":"train2.csv","type":"婚恋交友","range":range(867,977)},
        {"fileName":"more.csv","type":"Unknown","range":range(0,20642)},

        # {"fileName":"more_haveContent0402.csv","type":"Unknown","range":range(0,100)},
        # {"fileName":"more_haveContent0405.csv","type":"Unknown","range":range(0,20000)},

        {"fileName":"more_haveContent0405.csv","type":"购物消费","range":range(0,9)},           #1      9           7
        {"fileName":"more_haveContent0405.csv","type":"婚恋交友","range":range(9,5434)},        #2      5425        8
        {"fileName":"more_haveContent0405.csv","type":"假冒身份","range":range(5434,5480)},     #3      46          9
        {"fileName":"more_haveContent0405.csv","type":"钓鱼网站","range":range(5480,5848)},     #4      368         10
        {"fileName":"more_haveContent0405.csv","type":"冒充公检法","range":range(5848,5850)},   #5      2           11
        {"fileName":"more_haveContent0405.csv","type":"平台诈骗","range":range(5850,7311)},     #6      1461        12
        {"fileName":"more_haveContent0405.csv","type":"招聘兼职","range":range(7311,7319)},     #7      8           13
        {"fileName":"more_haveContent0405.csv","type":"杀猪盘","range":range(7319,7483)},       #8      164         14
        {"fileName":"more_haveContent0405.csv","type":"博彩赌博","range":range(7483,7719)},     #9      236         15
        {"fileName":"more_haveContent0405.csv","type":"信贷理财","range":range(7719,7738)},     #10     19          16
        {"fileName":"more_haveContent0405.csv","type":"刷单诈骗","range":range(7738,7761)},     #11     23          17
        {"fileName":"more_haveContent0405.csv","type":"中奖诈骗","range":range(7761,7764)},      #12     3          18


        {"fileName":"more_supplement2_seq_04013.csv","type":"购物消费","range":range(0,2)},     #1      2        19
        {"fileName":"more_supplement2_seq_04013.csv","type":"婚恋交友","range":range(2,434)},   #2      432      20
        {"fileName":"more_supplement2_seq_04013.csv","type":"冒充公检法","range":range(434,508)},  #5     74      21
        {"fileName":"more_supplement2_seq_04013.csv","type":"平台诈骗","range":range(508,916)},    #6     408       22
        {"fileName":"more_supplement2_seq_04013.csv","type":"博彩赌博","range":range(916,1218)},    #9      302     23
        {"fileName":"more_supplement2_seq_04013.csv","type":"信贷理财","range":range(1218,1292)},   #10     74      24
        {"fileName":"more_supplement2_seq_04013.csv","type":"刷单诈骗","range":range(1292,1456)},   #11     164     25

        #测试集
        #{"fileName":"train1.csv","type":None,"range":range(0.200)}
    ]

    allDataFeatureList:list = \
    [                                                                                           #label   len     idx
        {"fileName":"all_content.csv","type":"婚恋交友","range":range(2,33723)},                #2      33723   0
        {"fileName":"all_content.csv","type":"假冒身份","range":range(33723,33872)},            #3      149     1
        {"fileName":"all_content.csv","type":"钓鱼网站","range":range(33872,34342)},            #4      470     2
        {"fileName":"all_content.csv","type":"冒充公检法","range":range(34342,34490)},          #5      148     3
        {"fileName":"all_content.csv","type":"平台诈骗","range":range(34490,45005)},            #6      10515   4
        {"fileName":"all_content.csv","type":"招聘兼职","range":range(45005,45056)},            #7      51      5
        {"fileName":"all_content.csv","type":"杀猪盘","range":range(45056,45145)},              #8      89      6
        {"fileName":"all_content.csv","type":"博彩赌博","range":range(45145,46479)},            #9      1334    7
        {"fileName":"all_content.csv","type":"信贷理财","range":range(46479,50281)},            #10     3802    8
        {"fileName":"all_content.csv","type":"刷单诈骗","range":range(50281,52125)},            #11     1844    9
        {"fileName":"all_content.csv","type":"中奖诈骗","range":range(52125,52126)}             #12     1       10
    ]

    all_without_split_dataFeatureList:list = \
    [
        {"fileName":"all_content.csv","type":"购物消费","range":range(0,2)},                    #1      2       0
        {"fileName":"all_content.csv","type":"婚恋交友","range":range(2,6901)},                 #2      6899    1
        {"fileName":"all_content.csv","type":"假冒身份","range":range(6901,6937)},              #3      36      2
        {"fileName":"all_content.csv","type":"钓鱼网站","range":range(6937,7113)},              #4      176     3
        {"fileName":"all_content.csv","type":"冒充公检法","range":range(7113,7174)},            #5      61      4
        {"fileName":"all_content.csv","type":"平台诈骗","range":range(7174,8649)},              #6      1475    5
        {"fileName":"all_content.csv","type":"招聘兼职","range":range(8649,8657)},              #7      8       6
        {"fileName":"all_content.csv","type":"杀猪盘","range":range(8657,8757)},                #8      100     7
        {"fileName":"all_content.csv","type":"博彩赌博","range":range(8757,9095)},              #9      338     8
        {"fileName":"all_content.csv","type":"信贷理财","range":range(9095,10081)},             #10     986     9
        {"fileName":"all_content.csv","type":"刷单诈骗","range":range(10081,10662)},            #11     581     10
        {"fileName":"all_content.csv","type":"中奖诈骗","range":range(10662,10664)}             #12     2       11
    ]

    train_split_dataFeature:dict = \
    {
        "购物消费":{"filename":"all_content_split_train.csv", "range":range(0,2)},            
        "婚恋交友":{"filename":"all_content_split_train.csv","range":range(2,41919)},         #41917  
        "假冒身份":{"filename":"all_content_split_train.csv", "range":range(41919,42117)},     #200
        "钓鱼网站":{"filename":"all_content_split_train.csv", "range":range(42117,42664)},     #547
        "冒充公检法":{"filename":"all_content_split_train.csv","range":range(42664,42872)},   #208
        "平台诈骗":{"filename":"all_content_split_train.csv","range":range(42872,56297)},     #13425
        "招聘兼职":{"filename":"all_content_split_train.csv","range":range(56297,56321)},     #24
        "杀猪盘":{"filename":"all_content_split_train.csv","range":range(56321,56427)},       #106
        "博彩赌博":{"filename":"all_content_split_train.csv", "range":range(56427,58248)},     #1821
        "信贷理财":{"filename":"all_content_split_train.csv","range":range(58248,63208)},     #4960
        "刷单诈骗":{"filename":"all_content_split_train.csv","range":range(63208,65315)}      #2107
    }

    test_split_dataFeature:dict = \
    {
        "婚恋交友":{"filename":"all_content_split_train.csv", "range":range(0,11348)},    #11348
        "假冒身份":{"filename":"all_content_split_train.csv", "range":range(11348,11387)},    #39
        "钓鱼网站":{"filename":"all_content_split_train.csv", "range":range(11387,11602)},    #215
        "冒充公检法":{"filename":"all_content_split_train.csv", "range":range(11602,11647)},    #45
        "平台诈骗":{"filename":"all_content_split_train.csv", "range":range(11647,14182)},    #2535
        "招聘兼职":{"filename":"all_content_split_train.csv", "range":range(14182,14241)},    #59
        "杀猪盘":{"filename":"all_content_split_train.csv", "range":range(14241,14271)},    #30
        "博彩赌博":{"filename":"all_content_split_train.csv", "range":range(14271,14526)},    #255
        "信贷理财":{"filename":"all_content_split_train.csv", "range":range(14526,15713)},    #1187
        "刷单诈骗":{"filename":"all_content_split_train.csv", "range":range(15713,16501)}    #788
    }


 
def read_csv_context(filename:str,row_range:range,col:int = 1,decode:str='utf-8') -> list:
    """读取csv文件并依据指定的行数范围和列数将文本装载在列表中返回
        注意由于数据集的csv文件网页快照文本往往在第2列，
        col参数默认为第二列

    Args:
        filename (str): csv文件的路径
        row_range (range): 行数范围
        col (int): 列数，单个数据
        decode (str): 编码方式

    Returns:
        list: 装载文本的列表

    Example:
        read_csv_context("xx.csv",range(1,4),1)
        return:csv文件的第2,3,4排的第2列将作为数组的元素
        ["row2 col2","row3 col2","row4 col2"]
        
    Info:
        Created by LGD on 2023-3-9
        update 2023-4-1 可自定义字符编码方式
        Last update on 2023-3-9
    
    """
    data = []
    cnt = 0
    print("filename:%s col:%d",filename,col)
    with open(filename,encoding=decode) as csvfile:
        csv_reader = csv.reader(csvfile)
        for row in tqdm(csv_reader):
            if cnt in row_range:
                data.append(row[col])
            cnt += 1
    return data

def read_xlsx_context(filename:str,row_range:range,col:int) -> list:
    """读取csv文件并依据指定的行数范围和列数将文本装载在列表中返回

    Args:
        filename (str): csv文件的路径
        row_range (range):  行数范围
        col (int): 列数,由于数据集的数据排布是纵向，该方法依据一个范围的排读取指定列

    Returns:
        list: 装载文本的列表

    Example:
        read_xlsx_context("xx.xlsx",range(1,4),1)
        return:csv文件的第2,3,4排的第2列将作为数组的元素
        ["row2 col2","row3 col2","row4 col2"]
    
    Info:
        Created by LGD on 2023-3-9
        Last update on 2023-3-9
    
    """

    mapping_dict = {0:'A',1:'B',2:'C'}
    ws = load_workbook(filename)
    data:list = []

    print("filename:%s col:%d",filename,col)
    cnt:int = 0
    for cell in tqdm(ws["Sheet1"][mapping_dict[col]]):
        if cnt in row_range:
            data.append(cell.value)
        cnt += 1

    return data


def list_2_ndarray(int_lst:list) -> np.ndarray:
    """将整数列表转为numpy数组

    Args:
        int_lst (list): 整数列表

    Returns:
        np.ndarray: numpy数组

    Info:
        Created by LGD on 2023-3-19
    """
    return np.array(int_lst)

def split_word(nlp:spacy.Language,sentence:str,split_word:str = ' ') -> str:
    """将一个句子以split_word为间隔进行分词

    Args:
        nlp (spacy.Language): spacy Language模型
        sentence (str): 未分词的中文
        split_word (str, optional): 分割词符，默认为空格. Defaults to ' '.

    Returns:
        str: 分词后的字符串

    Example:
        split_word(nlp,"唤醒自己沉睡的财产房产")
        return:"唤醒 自己 沉睡 的 财产 房产"

    Info:
        Created by LGD on 2023-3-9
        Last update on 2023-3-9
    """
    words = nlp(sentence)
    lst = [str(word) for word in words]
    #word 的数据类型为 <class 'spacy.tokens.token.Token'>

    return split_word.join(word for word in lst)


def split_word_sentence_to_split_word_list(sentence:str) -> list:
    """将分词的句子转为单词的列表

    Args:
        sentence (str): 分词的字符串

    Returns:
        list: 该字符串所有单词的列表

    Example:
        split_word_sentence_to_split_word_list("唤醒 自己 沉睡 的 财产 房产")
        return:["唤醒" ,"自己" ,"沉睡" ,"的" ,"财产" ,"房产"]

    Info:
        Created by LGD on 2023-3-15
        Last update on 2023-3-15
    """
    return sentence.split(' ')


def split_word_from_sentence_array(nlp:spacy.Language,sentence_arr:list,split_word:str = ' ') -> list:
    """将一个句子以split_word为间隔进行分词

    Args:
        nlp (spacy.Language): spacy Language模型
        sentence (list): 未分词的中文
        split_word (str, optional): 分割词符，默认为空格. Defaults to ' '.

    Returns:
        str: 分词后的字符串

    Example:
        split_word_arr(nlp,["唤醒自己沉睡的财产房产","骗子竟是东方基金"])
        return:["唤醒 自己 沉睡 的 财产 房产","骗子 竟是 东方 基金"]

    Info:
        Created by LGD on 2023-3-9
        2023-4-10 调整了函数命名
        Last update on 2023-4-10
    """
    sentence_list = []
    for sentence in sentence_arr:
        words = nlp(sentence)
        single_sentence_lst = [str(word) for word in words]
        sentence_list.append(split_word.join(word for word in single_sentence_lst))
        
    return sentence_list

def split_to_word_set_from_sentence(nlp:spacy.Language,sentences:'str|list',throwStopWord:bool = False) -> 'list[list[str]]|list[str]':
    """将一个句子列表以单词为单位分割为单词的列表

    Args:
        nlp (spacy.Language): spacy Language模型
        sentence_arr (list): 未分词的中文

    Returns:
        list: 分词后的列表

    Example:
        split_word_arr(nlp,["唤醒自己沉睡的财产房产","骗子竟是东方基金"])
        return:[["唤醒" ,"自己" ,"沉睡", "的" ,"财产" ,"房产"],["骗子","竟是","东方" ,"基金"]]

    Info:
        Created by LGD on 2023-3-16
        2023-4-10 调整了函数命名和函数的参数类型提示
        Last update on 2023-4-10
        
    """
    print("分词中")
    if(type(sentences) == list):
        word_gather_list = []
        for sentence in tqdm(sentences):
            words = nlp(sentence)
            word_gather_list.append([str(word) for word in words])
        return word_gather_list
    elif(type(sentences) == str):
        words = nlp(sentences)
        word_gather = [str(word) for word in words]
        return word_gather


def normalization_word_number(wordSet:'list[list[str]]|list[str]',labelList:list = [],number:int=0) ->list:
    """归一化单词数量，当数量小于number时会被剔除，大于number时会归一到number的数量

    Args:
        wordList (list): 需要归一化的单词集，为单词集合的列表
        labelList (list):根据单词集的剔除情况

    Returns:
        list: 归一化后的列表

    Info:
        Created by LGD on 2023-3-19
    """
    if(type(wordSet[0]) == list):
        i:int = 0
        cnt:int = 0
        ori_len:int = len(wordSet)
        while i < ori_len - cnt:
            if len(wordSet[i]) < number:
                wordSet.pop(i)
                if(len(labelList)!=0):
                    labelList.pop(i)
                cnt += 1
                i -= 1
            i+=1

        wordSet = [word_set[0:number] for word_set in wordSet]
        return wordSet,labelList
    
    elif(type(wordSet[0]) == str):
        if(len(wordSet) < number):
            return []
        else:
            return wordSet[0:number] 

def throw_lower_than_threshold(wordSet:'list[list[str]]|list[str]',labelList:list = [],thresholdWordNum:int=0) ->list:
    """归一化单词数量，当数量小于specifiedWordNum时会重复填充，大于specifiedWordNum时会归一到specifiedWordNum的数量

    Args:
        wordList (list): 需要归一化的单词集，为单词集合的列表
        labelList (list):根据单词集的剔除情况
        specifiedWordNum (int):归一化数量

    Returns:
        list: 归一化后的列表

    Info:
        Created by LGD on 2023-3-19
    """
    if(type(wordSet[0]) == list):
        i:int = 0
        cnt:int = 0
        ori_len:int = len(wordSet)
        while i < ori_len - cnt:
            if len(wordSet[i]) <= thresholdWordNum:
                wordSet.pop(i)
                if(len(labelList)!=0):
                    labelList.pop(i)
                cnt += 1
                i -= 1
            print("上一次数量",len(wordSet[i]))
            i+=1

        return wordSet,labelList
    
    elif(type(wordSet[0]) == str):
        if len(wordSet[i]) <= thresholdWordNum:
            return "",[]
  
def new_normalization_word_number(wordSet:'list[list[str]]|list[str]',labelList:list = [],specifiedWordNum:int=0,thresholdWordNum:int=0) ->list:
    """归一化单词数量，当数量小于specifiedWordNum时会重复填充，大于specifiedWordNum时会归一到specifiedWordNum的数量

    Args:
        wordList (list): 需要归一化的单词集，为单词集合的列表
        labelList (list):根据单词集的剔除情况
        specifiedWordNum (int):归一化数量

    Returns:
        list: 归一化后的列表

    Info:
        Created by LGD on 2023-3-19
    """
    if(type(wordSet[0]) == list):
        i:int = 0
        cnt:int = 0
        ori_len:int = len(wordSet)
        while i < ori_len - cnt:
            if len(wordSet[i]) < specifiedWordNum and len(wordSet[i]) > thresholdWordNum:
                #将不足的部分进行重复
                while(len(wordSet[i]) < specifiedWordNum):
                    wordSet[i].extend(wordSet[i][0:specifiedWordNum-len(wordSet[i])])
            elif len(wordSet[i]) >= specifiedWordNum:
                wordSet[i] = wordSet[i][0:specifiedWordNum]
            else:
                wordSet.pop(i)
                if(len(labelList)!=0):
                    labelList.pop(i)
                cnt += 1
                i -= 1
            #print("上一次数量",len(wordSet[i]))
            if(len(wordSet[i])!= specifiedWordNum):
                pass
            i+=1

        return wordSet,labelList
    
    elif(type(wordSet[0]) == str):
        if(len(wordSet) < specifiedWordNum):
            while(len(wordSet) < specifiedWordNum):
                wordSet.extend(wordSet[0:specifiedWordNum-len(wordSet)])
            return wordSet
        else:
            return wordSet[0:specifiedWordNum]
        

        
def word_set_throw_stop_word(wordSet:'list[str]|list[list[str]]',stopWord:list) ->'list[list[str]]|list[str]':
    """依据停用词列表去除停用词

    Args:
        sentence_arr (list): 单词集列表
        stopWord (list): 停用词列表

    Returns:
        list: 

    Example:
        func([['天气','好'],['你','好']],['好'])
        return:[['天气'],['你']]

    Info:
        2023-4-10 调整了函数命名和函数的参数类型提示
        last update on 2023-4-10
    """
    if(type(wordSet[0]) == list):
        new_word_gather_list = []
        for word_gather in wordSet:
            new_word_gather = [word for word in word_gather if word not in stopWord]
            new_word_gather_list.append(new_word_gather)
        return new_word_gather_list
    elif(type(wordSet[0]) == str):
        new_word_gather = [word for word in wordSet if word not in stopWord]
        return new_word_gather



def wordGatherList_to_VectorList(nlp:spacy.Language,wordGatherList:list,is_flat:bool=False) -> list:
    """将单词集合的列表转换为词向量的列表

    Args:
        nlp (spacy.Language): spacy Language模型
        wordGatherList (list): 单词集合列表

    Returns:
        list: 词向量的列表

    Example:
        wordGatherList_to_VectorList([['cat','dog'],['red','blue']])
        result: [[333,222,444],[555,333,444]]  每个元素为输入的每个单词集合的词向量的拼接

    Info:
        Created by LGD on 2023-3-15
        Last update on 2023-3-15

    Update:
        2023-3-16 将输出结果由词向量的扁平拼接改为（单词数,300)的矩阵
    """
    embedding = []
    embedding_list = []
    for word_gather in wordGatherList:
        for word in word_gather:
            embedding:np.ndarray = np.append(embedding,nlp.vocab[word].vector)
        embedding_list.append(embedding)
        embedding = []

    new_embedding_list = []
    for i in range(len(wordGatherList)):
        embedding = embedding_list[i].reshape(len(wordGatherList[i]),-1)
        new_embedding_list.append(embedding)
        

    return new_embedding_list

def wordSet_to_Matrix(nlp:spacy.Language,wordSet:'list[list[str]]|list[str]',is_flat:bool=False) -> np.ndarray:
    """将单词集合的列表转换为词向量拼接的二维矩阵
        注意，由于python列表的元素长度可以任意，
            但二维ndarray的列数不可以，要求wordGather的长度必须归一化

    Args:
        nlp (spacy.Language): spacy Language模型
        wordGatherList (list): 单词集合列表

    Returns:
        list: 词向量的列表

    Example:
        wordGatherList_to_Matrix([['cat','dog'],['red','blue']])
        result: [[333,222,444],[555,333,444]]  每个元素为输入的每个单词集合的词向量的拼接

    Info:
        Created by LGD on 2023-3-15
        Last update on 2023-3-15

    Update:
        2023-4-10 调整了函数命名和函数的参数类型提示
        2023-3-16 将输出结果由词向量的扁平拼接改为（单词数,300)的矩阵
        Last update on 2023-4-10
    """
    cnt = 0
    if(type(wordSet[0]) == list):
        embedding = []
        for word_gather in wordSet:
            for word in word_gather:
                embedding:np.ndarray = np.append(embedding,nlp.vocab[word].vector)
            print("转换文本%d" %(cnt))
            cnt += 1
        embedding:np.ndarray = embedding.reshape(len(wordSet),-1)

        if(not is_flat):
            embedding = embedding.reshape(len(wordSet),len(wordSet[0]),300)

        return embedding
    elif(type(wordSet[0]) == str):
        embedding = []
        for word in wordSet:
            embedding:np.ndarray = np.append(embedding,nlp.vocab[word].vector)
        embedding:np.ndarray = embedding.reshape(len(wordSet),-1)
        if(not is_flat):
            embedding = embedding.reshape(1,len(wordSet),300)
        return embedding


def lower_dimension(array:np.ndarray) -> np.ndarray:
    """对高维向量进行降维
       似乎可以将任一的高维矩阵降成2维        
    
    Args:
        array (np.ndarray): 高维向量

    Returns:
        np.ndarray: 降维向量

    Info:
        Created by LGD on 2023-3-16
        Last update on 2023-3-16
    """
    tsne = TSNE()
    return tsne.fit_transform(array)


#TODO 尚未完成
def count_word_freq(word_gather:list) -> dict:
    """获取一个单词的集合（列表），将返回一个记录了所有单词的出现次数的字典

    Args:
        word_gather (list): 单词集合

    Returns:
        dict: 词频字典

    Example:
        count_word_freq(["cat","cat","mouse"])
        return:{"cat":2,"mouse":1}

    Info:
        Created by LGD on 2023-3-16
        Last update on 2023-3-16
    """
    pass


def get_similar_word(word:str) -> str:
    """获取最接近的单词

    Args:
        word (str): 待评估单词

    Returns:
        str: 最接近的单词
    """

def expand_content(wordSet:'list[list[str]]|list[str]',labelList:list = [],specifiedWordNum:int=0) ->list:
    if(type(wordSet[0]) == list):
        i:int = 0
        cnt:int = 0
        ori_len:int = len(wordSet)
        while i < ori_len + cnt:
            while(len(wordSet[i]) >= 2*specifiedWordNum):
                wordSet.insert(i,wordSet[i][0:specifiedWordNum])
                labelList.insert(i,labelList[i])
                i+=1
                cnt+=1
                wordSet[i] = wordSet[i][specifiedWordNum:]
            i+=1
        print("扩充了%d条数据" %(cnt))
        for wordset in wordSet:
            print(len(wordset))
    return wordSet,labelList

def label_reflect2class(label_list:'list[int]',n_class:int) -> 'list[int]':
    """将标签映射为[0-n_class-1]间的连续标签

    Args:
        label_list (list[int]): _description_
        n_class (int): _description_

    Returns:
        list[int]: _description_
    """
    reflect_label_dict = {}
    cnt:int = 0
    for label in label_list:
        if label not in reflect_label_dict:
            reflect_label_dict[label] = cnt
            cnt += 1
            if cnt > n_class:
                return []
    
    # 映射表创建完成
    new_label_list = [reflect_label_dict[label] for label in label_list]
    return new_label_list


def random_pick(texts:'list[str]',labels:'list[int] | list[str]',proportion:float):
    #随机抽选
    import pandas as pd
    data = \
    {
        "text":texts,
        "label":labels
    }
    line_label = [str(i) for i in range(len(labels))]
    df = pd.DataFrame(data,index=line_label)
    #交换部分帧
    sample = df.sample(n=int(proportion*len(labels)), frac=None, replace=False, weights=None,
            random_state=1,axis=0)
    return sample["text"].tolist(),sample["label"].tolist() 